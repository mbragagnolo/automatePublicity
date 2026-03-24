#!/usr/bin/env python3
"""automatePublicity — social media content generation pipeline.

Usage examples:
    python main.py "5 tips for better sleep" video --tiktok
    python main.py "Morning routine" video --reels --images photo1.jpg photo2.jpg
    python main.py "Quick Python trick" video --shorts
    python main.py "Before/after skincare" carousel --carousel-type before-after --before-image before.jpg --after-image after.jpg
    python main.py "Productivity hacks" video --tiktok --french
    python main.py "Fitness tips" video --reels --output-dir 2026-03-24
"""

import argparse
import importlib.util
import re
import sys
import tempfile
import types
from pathlib import Path
from datetime import date

ROOT = Path(__file__).parent
SUBPKGS = ROOT / "subpackages"

EXCEL_PATH = SUBPKGS / "Social-media-researcher" / "data" / "videos.xlsx"
CHROMA_DIR = ROOT / "data" / "chroma_db"

# ---------------------------------------------------------------------------
# Subpackage loader helpers
# ---------------------------------------------------------------------------

def _load_module(name: str, file_path: Path) -> types.ModuleType:
    """Load a single .py file as a named module and register it in sys.modules."""
    if name in sys.modules:
        return sys.modules[name]
    spec = importlib.util.spec_from_file_location(name, file_path)
    mod = importlib.util.module_from_spec(spec)
    sys.modules[name] = mod
    spec.loader.exec_module(mod)  # type: ignore[union-attr]
    return mod


def _load_package(name: str, pkg_dir: Path) -> types.ModuleType:
    """Load a Python package from an explicit directory and register it in sys.modules.

    Sets submodule_search_locations so that intra-package absolute imports
    (e.g. ``from social_media_expert.config import …``) resolve correctly
    without modifying sys.path.
    """
    if name in sys.modules:
        return sys.modules[name]
    init = pkg_dir / "__init__.py"
    spec = importlib.util.spec_from_file_location(
        name,
        init,
        submodule_search_locations=[str(pkg_dir)],
    )
    mod = importlib.util.module_from_spec(spec)
    sys.modules[name] = mod
    spec.loader.exec_module(mod)  # type: ignore[union-attr]
    return mod


# ---------------------------------------------------------------------------
# Load subpackages directly from their paths inside subpackages/
# ---------------------------------------------------------------------------

# English2french-translator — single file, loaded directly by path
_translator = _load_module(
    "translator",
    SUBPKGS / "English2french-translator" / "translator.py",
)

# photoshopConnector — Python package
_photoshopconnector = _load_package(
    "photoshopconnector",
    SUBPKGS / "photoshopConnector" / "photoshopconnector",
)

# socialmediaExpert — Python package (src layout)
_social_media_expert = _load_package(
    "social_media_expert",
    SUBPKGS / "socialmediaExpert" / "src" / "social_media_expert",
)

# ---------------------------------------------------------------------------
# Platform configuration
# ---------------------------------------------------------------------------
# Maps CLI platform flags to the platforms socialmediaExpert understands
PLATFORM_TO_SME = {
    "tiktok": "instagram",
    "reels": "instagram",
    "shorts": "instagram",
}

# Extra instructions injected into the LLM prompt to shape video script output
VIDEO_INSTRUCTIONS = {
    "tiktok": (
        "Write this as a TikTok video script. "
        "Structure: (1) Hook — one punchy sentence that grabs attention in the first 3 seconds; "
        "(2) Main content — 3 to 5 fast, energetic points with a casual tone; "
        "(3) CTA — tell the viewer to like, follow, or comment. "
        "Keep it under 60 seconds when read aloud. Label each section clearly."
    ),
    "reels": (
        "Write this as an Instagram Reels video script. "
        "Structure: (1) Hook — one eye-catching opening line for the first 3 seconds; "
        "(2) Main content — 3 to 5 short points with visual cues in [brackets]; "
        "(3) CTA — ask viewers to save, share, or follow. "
        "Keep it under 30 seconds when read aloud. Label each section clearly."
    ),
    "shorts": (
        "Write this as a YouTube Shorts video script. "
        "Structure: (1) Hook — one direct question or bold statement; "
        "(2) Main content — clear explanation in 3 to 5 steps; "
        "(3) CTA — ask viewers to subscribe and like. "
        "Keep it under 60 seconds when read aloud. Label each section clearly."
    ),
}


# ---------------------------------------------------------------------------
# Output directory helpers
# ---------------------------------------------------------------------------

def _next_run_number(date_dir: Path, platform: str) -> int:
    """Return the next available run number for a platform inside a date folder."""
    pattern = re.compile(rf"^{re.escape(platform)}_(\d+)$")
    existing = [
        int(m.group(1))
        for d in date_dir.iterdir()
        if d.is_dir() and (m := pattern.match(d.name))
    ] if date_dir.exists() else []
    return max(existing, default=0) + 1


def resolve_output_dir(output_dir_arg: str | None, platform: str) -> tuple[Path, str]:
    """Return (run_dir, run_name) where run_dir is <date>/<platform>_<n>/.

    The run number auto-increments so concurrent generations never collide.
    """
    if output_dir_arg:
        date_dir = Path(output_dir_arg)
        if not date_dir.is_absolute():
            date_dir = ROOT / date_dir
    else:
        date_dir = ROOT / date.today().strftime("%Y-%m-%d")

    run_n = _next_run_number(date_dir, platform)
    run_name = f"{platform}_{run_n}"
    return date_dir / run_name, run_name


def ensure_dirs(base: Path, platform: str, content_type: str) -> dict[str, Path]:
    """Create content-appropriate subdirectories and return them as a name→Path dict.

    Reels  : video/  cover_photo/  texts/
    TikTok/Shorts : video/  texts/
    Carousel : images/  texts/
    """
    if content_type == "carousel":
        names = ["images", "texts"]
    elif platform == "reels":
        names = ["video", "cover_photo", "texts"]
    else:  # tiktok, shorts
        names = ["video", "texts"]

    dirs = {n: base / n for n in names}
    for d in dirs.values():
        d.mkdir(parents=True, exist_ok=True)
    return dirs


# ---------------------------------------------------------------------------
# RAG: load transcripts from Social-media-researcher Excel and ingest into SME
# ---------------------------------------------------------------------------

def load_transcripts(excel_path: Path) -> list[str]:
    """Read the 'transcript' column from the Social-media-researcher Excel file."""
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        col_idx = headers.index("transcript")
    except ValueError:
        wb.close()
        return []
    transcripts = [
        str(row[col_idx]).strip()
        for row in ws.iter_rows(min_row=2, values_only=True)
        if col_idx < len(row) and row[col_idx] and str(row[col_idx]).strip()
    ]
    wb.close()
    return transcripts


def write_transcripts_pdf(transcripts: list[str], dest: Path) -> None:
    """Serialize transcript list to a simple PDF for RAG ingestion."""
    from fpdf import FPDF

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Social Media Video Transcripts", ln=True)
    pdf.ln(4)
    for i, transcript in enumerate(transcripts, 1):
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 8, f"Video {i}", ln=True)
        pdf.set_font("Helvetica", size=10)
        # FPDF requires latin-1 safe text; replace characters it can't encode
        safe = transcript.encode("latin-1", errors="replace").decode("latin-1")
        pdf.multi_cell(0, 6, safe)
        pdf.ln(3)
    pdf.output(str(dest))


def ingest_transcripts(expert, excel_path: Path) -> bool:
    """Read transcripts from Excel, write to temp PDF, ingest into expert. Returns True on success."""
    transcripts = load_transcripts(excel_path)
    if not transcripts:
        print("  No transcripts found in the Excel file.")
        return False

    print(f"  Found {len(transcripts)} transcript(s). Writing to temporary PDF...")
    tmp_pdf = Path(tempfile.mktemp(suffix=".pdf"))
    try:
        write_transcripts_pdf(transcripts, tmp_pdf)
        count = expert.ingest_file(tmp_pdf, metadata={"source_type": "transcript"})
        print(f"  Ingested {count} chunk(s) into knowledge base.")
        return True
    except ImportError:
        print("  WARNING: fpdf2 not installed — skipping transcript ingestion.")
        print("  Install it with: pip install fpdf2")
        return False
    except Exception as exc:
        print(f"  WARNING: Transcript ingestion failed: {exc}")
        return False
    finally:
        if tmp_pdf.exists():
            tmp_pdf.unlink()


# ---------------------------------------------------------------------------
# Content generation
# ---------------------------------------------------------------------------

def build_expert(use_rag: bool):
    """Instantiate SocialMediaExpert with project-level ChromaDB path."""
    return _social_media_expert.SocialMediaExpert(chroma_persist_dir=CHROMA_DIR)


def generate_content(
    expert,
    description: str,
    platform: str,
    extra_instructions: str,
    use_rag: bool,
) -> str:
    """Call socialmediaExpert and return the generated content string."""
    sme_platform = PLATFORM_TO_SME.get(platform, "instagram")
    post = expert.generate_post(
        topic=description,
        platform=sme_platform,
        extra_instructions=extra_instructions or None,
        use_rag=use_rag,
    )
    return post.content


def generate_caption(expert, description: str, content: str, platform: str, use_rag: bool) -> str:
    """Generate a social media caption from the post/slide content."""
    sme_platform = PLATFORM_TO_SME.get(platform, "instagram")
    context = content[:1500] if len(content) > 1500 else content
    instructions = (
        "Write an engaging social media post caption based on the content below. "
        "Include: a hook opening line, 2-3 sentences summarizing the key message, "
        "relevant emojis, a clear call to action, and 5-10 hashtags at the end. "
        "Write only the caption — no labels, no commentary.\n\n"
        f"Content:\n{context}"
    )
    post = expert.generate_post(
        topic=description,
        platform=sme_platform,
        extra_instructions=instructions,
        use_rag=use_rag,
    )
    return post.content


def generate_hook_variants(expert, description: str, platform: str, n: int, use_rag: bool) -> list[str]:
    """Generate n short hook lines for reel covers — one LLM call, parsed into a list."""
    sme_platform = PLATFORM_TO_SME.get(platform, "instagram")
    instructions = (
        f"Generate exactly {n} alternative title hooks for a short-form video cover/thumbnail "
        f"about this topic. Each hook must be 3 to 4 words maximum — ultra-short and punchy, "
        f"like a magazine cover headline. Return ONLY the hooks, one per line, no numbering, "
        f"no punctuation at the end, no extra text."
    )
    post = expert.generate_post(
        topic=description,
        platform=sme_platform,
        extra_instructions=instructions,
        use_rag=use_rag,
    )
    hooks = [line.strip() for line in post.content.splitlines() if line.strip()]
    # Pad with the first hook if the LLM returned fewer than requested
    while len(hooks) < n:
        hooks.append(hooks[0] if hooks else description)
    return hooks[:n]


# ---------------------------------------------------------------------------
# Translation
# ---------------------------------------------------------------------------

def translate_to_french(text: str) -> str:
    """Translate text to French via the English2french-translator submodule."""
    return _translator.translate(text)


# ---------------------------------------------------------------------------
# Photoshop asset generation
# ---------------------------------------------------------------------------

def generate_reel_covers(name: str, takes: list[tuple[str, str]], cover_dir: Path) -> list:
    """Generate one reel cover per (image_path, hook_text) pair.

    Args:
        name: Base name for output files.
        takes: List of (image_path, hook_text) — one entry per image/hook combination.
        cover_dir: Directory where covers are saved.

    Returns:
        List of (jpeg_path, psd_path) tuples, one per take.
    """
    results = []
    for i, (image_path, hook_text) in enumerate(takes, 1):
        jpeg, psd = _photoshopconnector.create_reel_cover(
            name=f"{name}_cover_{i}",
            text=hook_text,
            image_path=image_path,
            output_dir=str(cover_dir),
            auto_crop=True,
        )
        results.append((jpeg, psd))
    return results


def generate_carousel_slides(
    carousel_type: str,
    name: str,
    slide_texts: list[str],
    image: str | None,
    before_image: str | None,
    after_image: str | None,
    assets_dir: Path,
) -> list:
    """Generate carousel slides via photoshopConnector. Returns list of output paths."""
    if carousel_type == "repeated":
        return _photoshopconnector.create_carousel_repeated(
            name=name,
            texts=slide_texts,
            image_path=image,
            output_dir=str(assets_dir),
            auto_crop=True,
        )
    elif carousel_type == "panorama":
        return _photoshopconnector.create_carousel_panorama(
            name=name,
            image_path=image,
            n_slides=len(slide_texts),
            output_dir=str(assets_dir),
        )
    elif carousel_type == "before-after":
        return _photoshopconnector.create_carousel_before_after(
            name=name,
            treatment_text=name,
            before_image=before_image,
            after_image=after_image,
            output_dir=str(assets_dir),
            auto_crop=True,
        )
    else:
        raise ValueError(f"Unknown carousel type: {carousel_type}")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def extract_hook(content: str, max_len: int = 200) -> str:
    """Return the first paragraph/line of content, truncated, for use as cover text."""
    first = content.split("\n")[0].strip()
    return first[:max_len] if len(first) > max_len else first


def content_to_slide_texts(content: str, max_slides: int = 6) -> list[str]:
    """Split generated content into per-slide text chunks (one per paragraph)."""
    paragraphs = [p.strip() for p in content.split("\n\n") if p.strip()]
    if not paragraphs:
        paragraphs = [content.strip()]
    return paragraphs[:max_slides]


def parse_slide_text(raw: str) -> list[str]:
    """Parse '[Slide 1] text [Slide 2] text ...' into an ordered list of slide strings."""
    parts = re.split(r"\[Slide\s+\d+\]", raw, flags=re.IGNORECASE)
    return [p.strip() for p in parts if p.strip()]


def safe_name(description: str, max_len: int = 30) -> str:
    """Create a filesystem-safe name from the description string."""
    slug = re.sub(r"[^\w\s-]", "_", description[:max_len].lower())
    return re.sub(r"[\s_-]+", "_", slug).strip("_")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="main.py",
        description="automatePublicity — generate social media content automatically",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )

    parser.add_argument("description", help="Subject / brief for the content to generate")
    parser.add_argument(
        "content_type",
        choices=["video", "carousel"],
        help="Type of content: 'video' (script) or 'carousel' (slides)",
    )

    platform_group = parser.add_mutually_exclusive_group()
    platform_group.add_argument("--tiktok", action="store_true", help="Format for TikTok")
    platform_group.add_argument("--reels", action="store_true", help="Format for Instagram Reels (also generates covers)")
    platform_group.add_argument("--shorts", action="store_true", help="Format for YouTube Shorts")

    parser.add_argument("--french", action="store_true", help="Translate output to French")
    parser.add_argument("--output-dir", metavar="DIR", help="Output directory (default: today's date)")
    parser.add_argument("--name", metavar="NAME", help="Base name for output files (default: derived from description)")
    parser.add_argument("--skip-rag", action="store_true", help="Skip transcript ingestion from Social-media-researcher")

    # Image inputs for Photoshop steps
    parser.add_argument(
        "--images", metavar="PATH", nargs="+",
        help="One or more source images. For reels: one cover is generated per image "
             "(each with a different hook). For carousel: only the first image is used.",
    )
    parser.add_argument("--before-image", metavar="PATH", help="'Before' image for before-after carousel")
    parser.add_argument("--after-image", metavar="PATH", help="'After' image for before-after carousel")
    parser.add_argument(
        "--carousel-type",
        choices=["repeated", "panorama", "before-after"],
        default="repeated",
        help="Carousel layout type (default: repeated)",
    )
    parser.add_argument(
        "--slide-file",
        metavar="FILE",
        help=(
            "Path to a .txt file with slide content, bypassing AI generation (repeated carousel only). "
            "Format: [Slide 1] followed by the slide text, repeated for each slide."
        ),
    )

    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    # ------------------------------------------------------------------
    # Resolve platform
    # ------------------------------------------------------------------
    if args.content_type == "video":
        if args.reels:
            platform = "reels"
        elif args.shorts:
            platform = "shorts"
        else:
            platform = "tiktok"
            if not args.tiktok:
                print("No platform flag given — defaulting to --tiktok.")
    else:
        platform = "instagram"

    name = args.name or safe_name(args.description)

    # ------------------------------------------------------------------
    # Prepare output directories  (<date>/<platform>_<n>/)
    # ------------------------------------------------------------------
    output_dir, run_name = resolve_output_dir(args.output_dir, platform)
    dirs = ensure_dirs(output_dir, platform, args.content_type)
    print(f"Run: {run_name}  →  {output_dir}")

    # ------------------------------------------------------------------
    # Bypass mode: only for repeated carousel with --slide-text provided
    # ------------------------------------------------------------------
    if args.slide_file and args.content_type == "carousel" and args.carousel_type != "repeated":
        print(f"NOTE: --slide-text is only used with --carousel-type repeated; ignored for {args.carousel_type}.")
    bypass = args.content_type == "carousel" and args.carousel_type == "repeated" and bool(args.slide_file)

    # ------------------------------------------------------------------
    # [1/4] Instantiate socialmediaExpert (always needed — caption generation)
    # ------------------------------------------------------------------
    print("\n[1/4] Setting up AI expert...")
    expert = build_expert(use_rag=not args.skip_rag)

    # ------------------------------------------------------------------
    # [2/4] RAG: ingest transcripts from Social-media-researcher Excel
    # ------------------------------------------------------------------
    rag_active = False
    if bypass:
        print("\n[2/4] RAG skipped (slide file provided).")
    elif args.skip_rag:
        print("\n[2/4] RAG skipped (--skip-rag).")
    elif EXCEL_PATH.exists():
        print(f"\n[2/4] Loading research data from {EXCEL_PATH.name}...")
        rag_active = ingest_transcripts(expert, EXCEL_PATH)
    else:
        print(
            "\n[2/4] No research data found — skipping RAG enrichment.\n"
            "       Run Social-media-researcher first to build the knowledge base."
        )

    # ------------------------------------------------------------------
    # [3/4] Content generation
    # ------------------------------------------------------------------
    lang_suffix = "_fr" if args.french else ""

    if bypass:
        slide_file = Path(args.slide_file)
        if not slide_file.exists():
            print(f"ERROR: --slide-file not found: {slide_file}")
            sys.exit(1)
        print(f"\n[3/4] Reading slide file: {slide_file.name} (no AI generation).")
        raw = slide_file.read_text(encoding="utf-8")
        slide_texts = parse_slide_text(raw)
        print(f"      Parsed {len(slide_texts)} slide(s).")
        content = raw  # preserve original formatting in the saved .txt
    else:
        extra_instructions = VIDEO_INSTRUCTIONS.get(platform, "") if args.content_type == "video" else ""
        print(f"\n[3/4] Generating {args.content_type} content...")
        print(f"      Topic    : {args.description!r}")
        print(f"      Platform : {platform}")
        print(f"      RAG      : {'active' if rag_active else 'off'}")
        content = generate_content(expert, args.description, platform, extra_instructions, rag_active)

        if args.french:
            print("      Translating to French...")
            content = translate_to_french(content)

    # Save content to texts/ folder
    script_path = dirs["texts"] / f"{name}{lang_suffix}.txt"
    script_path.write_text(content, encoding="utf-8")
    print(f"\n      Content saved → {script_path}")

    # Generate and save caption
    print("      Generating post caption...")
    caption = generate_caption(expert, args.description, content, platform, rag_active)
    if args.french:
        caption = translate_to_french(caption)
    caption_path = dirs["texts"] / f"{name}{lang_suffix}_caption.txt"
    caption_path.write_text(caption, encoding="utf-8")
    print(f"      Caption saved  → {caption_path}")

    # Preview
    preview = content[:600] + ("\n..." if len(content) > 600 else "")
    print("\n--- Content Preview " + "-" * 40)
    print(preview)
    print("--- Caption Preview " + "-" * 40)
    print(caption[:400] + ("\n..." if len(caption) > 400 else ""))
    print("-" * 60)

    # ------------------------------------------------------------------
    # [4/4] Photoshop assets
    # ------------------------------------------------------------------
    print(f"\n[4/4] Generating visual assets...")

    if args.content_type == "video" and platform == "reels":
        images = args.images or []
        if not images:
            print("      NOTE: Pass --images <path> [<path> ...] to generate reel covers.")
        else:
            n = len(images)
            print(f"      Generating {n} hook variant(s) for {n} image(s)...")
            try:
                hooks = generate_hook_variants(expert, args.description, platform, n, rag_active)
                takes = list(zip(images, hooks))
                for img, hook in takes:
                    print(f"        Image: {img}  |  Hook: {hook!r}")
                covers = generate_reel_covers(name, takes, dirs["cover_photo"])
                for jpeg, psd in covers:
                    print(f"        Cover → {jpeg}")
            except Exception as exc:
                print(f"      WARNING: Reel cover generation failed: {exc}")

    elif args.content_type == "carousel":
        images = args.images or []
        needs_image = args.carousel_type in ("repeated", "panorama")
        needs_both = args.carousel_type == "before-after"

        if needs_image and not images:
            print(f"      NOTE: Pass --images <path> to generate a {args.carousel_type} carousel.")
        elif needs_both and (not args.before_image or not args.after_image):
            print("      NOTE: Pass --before-image and --after-image to generate a before-after carousel.")
        else:
            print(f"      Generating {args.carousel_type} carousel via Photoshop...")
            try:
                if not bypass:
                    slide_texts = content_to_slide_texts(content)
                slides = generate_carousel_slides(
                    carousel_type=args.carousel_type,
                    name=name,
                    slide_texts=slide_texts,
                    image=images[0] if images else None,
                    before_image=args.before_image,
                    after_image=args.after_image,
                    assets_dir=dirs["images"],
                )
                for slide in slides:
                    print(f"        Slide → {slide}")
            except Exception as exc:
                print(f"      WARNING: Carousel generation failed: {exc}")

    else:
        print("      No visual assets for this content type / platform combination.")

    print("\nDone.")


if __name__ == "__main__":
    main()
